import openpyxl

wb = openpyxl.load_workbook('/home/bittu/Desktop/MOM/Vertical and Entity name.xlsx', read_only=True)
print('Sheet names:', wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n=== Sheet: {sheet_name} ===')
    print(f'Rows: {ws.max_row}, Cols: {ws.max_column}')
    # Print header row
    headers = [cell.value for cell in ws[1]]
    print(f'Headers: {headers}')
    # Print first 5 data rows
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True), 2):
        print(f'  Row {i}: {row}')

wb.close()

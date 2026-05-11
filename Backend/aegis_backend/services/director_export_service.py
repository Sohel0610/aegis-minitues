import os
import io
import zipfile
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Styles
HEADER_FILL = PatternFill(start_color="75479C", end_color="75479C", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)
SUBHEADER_FILL = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
BORDER = Border(
    left=Side(style='thin', color="D1D5DB"),
    right=Side(style='thin', color="D1D5DB"),
    top=Side(style='thin', color="D1D5DB"),
    bottom=Side(style='thin', color="D1D5DB")
)

def format_sheet(ws):
    """Apply standard professional formatting to a sheet"""
    for row in ws.iter_rows():
        for cell in row:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if cell.row == 1:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    ws.row_dimensions[1].height = 35

def generate_director_excel(data):
    """
    Generates a formatted Excel workbook for a single director
    data: dict matching the output of director_full.py
    """
    wb = Workbook()
    
    # 1. Profile Sheet
    ws_profile = wb.active
    ws_profile.title = "Director Profile"
    
    profile_data = [
        ["Field", "Information"],
        ["DIN", data['director_info'].get('din')],
        ["Full Name", data['director_info'].get('name')],
        ["DIN Status", data['director_info'].get('din_status')],
        ["Gender", data['director_info'].get('gender')],
        ["Nationality", data['director_info'].get('nationality')],
        ["Date of Birth", data['profile'].get('date_of_birth')],
        ["PAN", data['profile'].get('pan')],
        ["Qualification", data['profile'].get('qualification')],
        ["Address", data['profile'].get('address')],
        ["Experience", data['profile'].get('experience')]
    ]
    
    for row in profile_data:
        ws_profile.append(row)
    
    # Column widths for Profile
    ws_profile.column_dimensions['A'].width = 25
    ws_profile.column_dimensions['B'].width = 80
    format_sheet(ws_profile)

    # 2. Associations Sheet
    ws_assoc = wb.create_sheet("Associations")
    ws_assoc.append(["S.No.", "CIN", "Company Name", "Designation", "Appt Date", "Status"])
    
    # Sort associations alphabetically by company name
    associations = data.get('associations', [])
    associations.sort(key=lambda x: (x.get('company_name') or "").lower())
    
    for i, assoc in enumerate(associations, 1):
        appt_date = assoc.get('appointment_date')
        status = assoc.get('company_status') or assoc.get('status') or "N/A"
        ws_assoc.append([
            i,
            assoc.get('cin') or "N/A",
            assoc.get('company_name') or "N/A",
            assoc.get('designation') or "Director",
            str(appt_date) if appt_date else "N/A",
            status
        ])
    
    ws_assoc.column_dimensions['A'].width = 8
    ws_assoc.column_dimensions['B'].width = 25
    ws_assoc.column_dimensions['C'].width = 50
    ws_assoc.column_dimensions['D'].width = 25
    ws_assoc.column_dimensions['E'].width = 15
    ws_assoc.column_dimensions['F'].width = 15
    format_sheet(ws_assoc)

    # 3. Family Sheet
    ws_family = wb.create_sheet("Family Information")
    ws_family.append(["S.No.", "Relationship", "Name", "PAN"])
    
    family_rows = []
    added_names = set() # To track normalized names for deduplication

    # A. New Relational Data (Priority - contains PANs)
    members = data.get('family_members', [])
    for m in members:
        name = (m.get('full_name') or m.get('details') or "").strip()
        if not name or name.lower() in ["nil", "none", "n/a"]:
            continue
            
        norm_name = " ".join(name.lower().split()) # Normalize whitespace
        if norm_name not in added_names:
            family_rows.append({
                "rel": m.get('relationship') or "Relative",
                "name": name,
                "pan": m.get('pan') or m.get('pan_number') or "N/A"
            })
            added_names.add(norm_name)
    
    # B. Master/Legacy Data (Fallback or Supplement)
    fam = data.get('family') or {}
    field_map = [
        ("Father", fam.get('father')),
        ("Mother", fam.get('mother')),
        ("Spouse", fam.get('section_2_77_ii')),
        ("Son", fam.get('son')),
        ("Son's Wife", fam.get('sons_wife')),
        ("Daughter", fam.get('daughter')),
        ("Daughter's Husband", fam.get('daughters_husband')),
        ("Brother", fam.get('brother')),
        ("Sister", fam.get('sister'))
    ]
    
    for rel, val in field_map:
        if val and str(val).lower() not in ["nil", "none", "n/a", ""]:
            # Handle multiple names in one string (commas or double spaces)
            names = []
            if "," in str(val):
                names = [n.strip() for n in str(val).split(",")]
            elif "  " in str(val): 
                names = [n.strip() for n in str(val).split("  ")]
            else:
                names = [str(val).strip()]
            
            for name in names:
                if not name: continue
                norm_name = " ".join(name.lower().split())
                if norm_name not in added_names:
                    family_rows.append({
                        "rel": rel,
                        "name": name,
                        "pan": "N/A" # Legacy columns don't have individual PANs in this loop
                    })
                    added_names.add(norm_name)
    
    if not family_rows:
        ws_family.append([1, "N/A", "No family records found", "N/A"])
    else:
        # Sort family members by relationship type for a professional look
        family_rows.sort(key=lambda x: x['rel'])
        for i, row in enumerate(family_rows, 1):
            ws_family.append([i, row['rel'], row['name'], row['pan']])
    
    # Done with family section

    ws_family.column_dimensions['A'].width = 8
    ws_family.column_dimensions['B'].width = 25
    ws_family.column_dimensions['C'].width = 50
    ws_family.column_dimensions['D'].width = 20
    format_sheet(ws_family)

    # Final touch: Row height for large text
    for ws in wb.worksheets:
        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = 30 if row > 1 else 35

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def create_bulk_zip(all_directors_data):
    """
    Creates a ZIP folder containing individual Excel files
    all_directors_data: list of dicts (each from director_full.py)
    """
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'a', zipfile.ZIP_DEFLATED, False) as zip_file:
        for data in all_directors_data:
            name = data['director_info'].get('name', 'Unknown').replace(" ", "_")
            din = data['director_info'].get('din', '00000000')
            filename = f"Director_Disclosure_{din}_{name}.xlsx"
            
            excel_buffer = generate_director_excel(data)
            zip_file.writestr(filename, excel_buffer.getvalue())
            
    zip_buffer.seek(0)
    return zip_buffer

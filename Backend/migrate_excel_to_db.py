import os
import sqlite3
import re
from openpyxl import load_workbook

def clean_vertical_name(name):
    if not isinstance(name, str) or not name.strip():
        return "Others"
    name = name.strip()
    lname = name.lower()
    if "realty" in lname:
        return "Realty"
    if "cement" in lname or "cemindia" in lname:
        return "Cement"
    if "airport" in lname:
        return "Airport"
    if "natural resources" in lname:
        return "Natural Resources"
    if "promoter" in lname:
        return "Promoter"
    if "thermal power" in lname:
        return "Thermal Power"
    if "renewable" in lname:
        return "Renewables"
    if "transmission" in lname:
        return "Transmission & Distribution"
    if "solar" in lname:
        return "Solar Manufacturing"
    if "gas" in lname:
        return "Gas Distribution"
    if "data centre" in lname or "datacenter" in lname:
        return "Data Centre"
    return name

def generate_code(name, existing_codes):
    words = re.findall(r'[A-Za-z0-9]+', name)
    if len(words) >= 3:
        code = (words[0][0] + words[1][0] + words[2][0]).upper()
    elif len(words) == 2:
        code = (words[0][:2] + words[1][0]).upper()
    elif len(words) == 1:
        code = words[0][:3].upper()
    else:
        code = "OTH"
    
    original = code
    counter = 1
    while code in existing_codes:
        code = f"{original[:2]}{counter}"
        counter += 1
    existing_codes.add(code)
    return code

def migrate():
    excel_path = os.path.abspath(r"d:\MOM\Vertical and Entity name.xlsx")
    db_path = os.path.abspath(r"d:\MOM\Backend\aegis_backend\public\local_fallback.db")
    
    print(f"Opening workbook: {excel_path}")
    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    print(f"Total rows read: {len(rows)}")

    # Row 0: ['Vertical Name', 'Name of Companies', 'Company Secretary' / name]
    # Header check
    data_rows = []
    for r in rows[2:]: # Skip header row and initial title row
        if not r or len(r) < 2: continue
        raw_v = r[0]
        comp_name = r[1]
        sec_name = r[2] if len(r) > 2 else None
        
        if raw_v and comp_name and str(comp_name).strip():
            data_rows.append((str(raw_v).strip(), str(comp_name).strip(), str(sec_name).strip() if sec_name else None))

    print(f"Valid data entries: {len(data_rows)}")

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("PRAGMA table_info(companies)")
    cols = [r['name'] for r in cursor.fetchall()]
    if 'secretary_name' not in cols:
        cursor.execute("ALTER TABLE companies ADD COLUMN secretary_name TEXT")

    # 1. Verticals
    raw_verticals = set(r[0] for r in data_rows)
    cleaned_verticals = sorted(list(set(clean_vertical_name(v) for v in raw_verticals)))
    
    existing_codes = set()
    cursor.execute("SELECT id, name, code FROM verticals")
    for r in cursor.fetchall():
        existing_codes.add(r['code'])

    vertical_id_map = {}

    for vert_name in cleaned_verticals:
        cursor.execute("SELECT id, code FROM verticals WHERE LOWER(name) = LOWER(?)", (vert_name,))
        row = cursor.fetchone()
        if row:
            v_id = row['id']
            vertical_id_map[vert_name] = v_id
        else:
            code = generate_code(vert_name, existing_codes)
            cursor.execute("INSERT INTO verticals (name, code) VALUES (?, ?)", (vert_name, code))
            v_id = cursor.lastrowid
            vertical_id_map[vert_name] = v_id
            print(f"  + Added Vertical: '{vert_name}' (Code: {code}, ID: {v_id})")

    # Fetch existing companies
    cursor.execute("SELECT id, name FROM companies")
    existing_companies = {r['name'].lower().strip(): r['id'] for r in cursor.fetchall()}

    added_companies = 0
    updated_companies = 0
    
    for raw_v, comp_name, sec_name in data_rows:
        vert_clean = clean_vertical_name(raw_v)
        v_id = vertical_id_map.get(vert_clean)

        comp_key = comp_name.lower()
        if comp_key in existing_companies:
            c_id = existing_companies[comp_key]
            cursor.execute("""
                UPDATE companies 
                SET vertical_id = ?, secretary_name = COALESCE(?, secretary_name)
                WHERE id = ?
            """, (v_id, sec_name, c_id))
            updated_companies += 1
        else:
            c_type = "Public" if "limited" in comp_name.lower() and "private" not in comp_name.lower() else "Private"
            cursor.execute("""
                INSERT INTO companies (name, type, vertical_id, status, secretary_name)
                VALUES (?, ?, ?, 'Active', ?)
            """, (comp_name, c_type, v_id, sec_name))
            new_id = cursor.lastrowid
            existing_companies[comp_key] = new_id
            added_companies += 1

    conn.commit()

    cursor.execute("SELECT COUNT(*) as count FROM verticals")
    v_total = cursor.fetchone()['count']
    cursor.execute("SELECT COUNT(*) as count FROM companies")
    c_total = cursor.fetchone()['count']
    cursor.execute("SELECT COUNT(*) as count FROM companies WHERE vertical_id IS NOT NULL")
    c_mapped = cursor.fetchone()['count']

    print(f"\n=== MIGRATION COMPLETE ===")
    print(f"Total Verticals in DB: {v_total}")
    print(f"Total Companies in DB: {c_total}")
    print(f"Companies linked to Verticals: {c_mapped}")
    print(f"New Companies added: {added_companies}")
    print(f"Existing Companies updated: {updated_companies}")

    conn.close()

if __name__ == "__main__":
    migrate()
